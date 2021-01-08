#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Author: Sierra Nguyen
Date:   10/11/2020
Title:  Sellics Advertising Audit Automation
"""

# change working directory
import os
os.chdir('/Users/Boo Boo/Downloads/Ad-Audit')

# importing openpyxl module 
import openpyxl as xl

# background info

email = input("Enter their account's email: ") 

'''
from tkinter import Tk, Label, Button, Canvas, Entry

email = []

master = Tk()
master.title('Account information')
master.resizable(False, False)
c = Canvas(master, width=500, height=150)
c.pack()

label = Label(master, text='What is the account\'s email address?')
label.pack()
label.place(relx=0.25, rely=0.2)

ent=Entry(master)
ent.pack()
ent.place(relx=0.305, rely=0.35)

btn1 = Button(master, text='Enter', command=lambda: email.append(ent.get()))
btn1.pack()
btn1.place(relx=0.45, rely=0.6)

btn2 = Button(master, text='Done', command=master.destroy)
btn2.pack()
btn2.place(relx=0.45, rely=0.75)

email = email[0]
'''
sep = '.'
email = email.split(sep, 1)[0]

# opening the source excel file 

# get source filenames
allfiles = os.listdir()

for value in allfiles:
    if 'Targets' in value:
        target = value
    elif 'Searchterms' in value:
        searchterm = value

# targets
tg = xl.load_workbook(filename=target) 
tg_sheet = tg.active

# searchterms
st = xl.load_workbook(filename=searchterm)
st_sheet = st.active

# opening the destination excel file and sheets
template = 'Advertising Audit Template_2020 - Final.xlsx'
main = xl.load_workbook(filename=template)
main.get_sheet_names()

main_tg = main['TGT Report']
main_st = main['ST Report']

# calculate total number of rows and columns in source excel file 

# targets
mr = tg_sheet.max_row 
mc = tg_sheet.max_column

# copying the cell values from source to destination
for i in range(1, mr+1):
    for j in range(1, mc+1):
        # reading cell value from source excel file 
        c = tg_sheet.cell(row = i, column = j)
        
        # writing the read value to destination excel file 
        main_tg.cell(row = i, column = j).value = c.value 
        
# searchterms
mr = st_sheet.max_row 
mc = st_sheet.max_column

# copying the cell values from source to destination
for i in range(1, mr+1):
    for j in range(1, mc+1):
        # reading cell value from source excel file 
        c = st_sheet.cell(row = i, column = j)
        
        # writing the read value to destination excel file 
        main_st.cell(row = i, column = j).value = c.value 

# saving the destination excel file 
main.save(str('AdAudit-' + email + '.xlsx'))

# removing input files

# targets
if os.path.exists(target):
    os.remove(target)
else:
    print("The file does not exist!")

# searchterms
if os.path.exists(searchterm):
  os.remove(searchterm)
else:
  print("The file does not exist!") 