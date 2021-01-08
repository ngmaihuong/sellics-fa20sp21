#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Author: Sierra Nguyen
Date:   12/04/2020
Title:  Sellics Value Statement Test Automation
"""

# change working directory
import os
os.chdir('/Users/Boo Boo/Downloads/Value-Statement')

# importing openpyxl module 
import openpyxl as xl

# background info
userid = input("Enter their account's ID: ") 

# opening the source excel file 

# get source filenames
allfiles = os.listdir()

for value in allfiles:
    if 'Targets' in value:
        target = value
    elif 'Change-Log' in value:
        changelog = value

# targets
tg = xl.load_workbook(filename=target) 
tg_sheet = tg.active

# change log
cl = xl.load_workbook(filename=changelog)
cl_sheet = cl.active

# opening the destination excel file and sheets
template = 'template-optimization.xlsx'
main = xl.load_workbook(filename=template)
main.get_sheet_names()

main_tg = main['Targets']
main_cl = main['Change Log']

# calculate total number of rows and columns in source excel file 

# targets
mr = tg_sheet.max_row 
mc = tg_sheet.max_column

# copying the cell values from source to destination
for i in range(1, mr+1):
    for j in range(1, mc+1):
        # reading cell value from source excel file 
        c = tg_sheet.cell(row = i, column = j)
        
        # writing the read value to destination excel file 
        main_tg.cell(row = i, column = j).value = c.value 
        
# searchterms
mr = cl_sheet.max_row 
mc = cl_sheet.max_column

# copying the cell values from source to destination
for i in range(1, mr+1):
    for j in range(1, mc+1):
        # reading cell value from source excel file 
        c = cl_sheet.cell(row = i, column = j)
        
        # writing the read value to destination excel file 
        main_cl.cell(row = i, column = j).value = c.value 

# saving the destination excel file 
main.save(str(userid + '.xlsx'))

# removing input files

# targets
if os.path.exists(target):
    os.remove(target)
else:
    print("The file does not exist!")

# searchterms
if os.path.exists(changelog):
  os.remove(changelog)
else:
  print("The file does not exist!") 